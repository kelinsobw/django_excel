from control.forms import Weighing, read_file, transliterate_def
import datetime
import openpyxl
import pathlib
from django import forms
from django.shortcuts import render, redirect
from openpyxl.styles import Font
import os
from django.http import HttpResponse


def history_save(text):
    f = open('history.txt', 'a')
    f.write(str(datetime.datetime.now()) + text + '\n')
    f.close()


def plus(request):
    class Plus(forms.Form):
        head = read_file('head')
        choice_room = []
        for el in range(len(head)):
            choice_room.append((str(transliterate_def(head[el])), str(head[el])))
        p_choice_room = forms.ChoiceField(choices=choice_room, label='', initial=None)
    cosmetic_in_cab, choice_room = build_vir_cab("Наименование")

    if request.method == "POST":
        form = Plus(request.POST, request.FILES)
        if form.is_valid():
            info_base = form.cleaned_data
            return redirect(f"/plus/{info_base.get('p_choice_room').lower()}/{str(transliterate_def(cosmetic_in_cab[0][0]))}")
    else:
        form = Plus()
        return render(request, "control/weighing.html", {"form": form})


def error_weidth():
    return HttpResponse("jkhgfds")


def build_vir_cab(cabinet, setting=0): #setting == 0=>don't price, 1>price
    data_now = read_file('all')
    head = read_file('head')
    cabinet_num = 0

    if cabinet == "Наименование":
        cosmetic_in_cab = []
        for i in range(1, len(data_now)):
            cosmetic_in_cab.append((data_now[i][2],data_now[i][3], data_now[i][4]))
        choice_room = []
        return (cosmetic_in_cab, choice_room)

    for i in range(5, len(data_now[0])):
        if transliterate_def(data_now[0][i]) == cabinet:
            cabinet_num = i
    cosmetic_in_cab = []
    for i in range(1, len(data_now)):
        if str(data_now[i][cabinet_num]) != "nan":
            if str(setting) == "1":
                cosmetic_in_cab.append((data_now[i][2], data_now[i][3], data_now[i][4], data_now[i][cabinet_num], data_now[i][6]/data_now[i][4]))
            else:
                cosmetic_in_cab.append((data_now[i][2], data_now[i][3], data_now[i][4], data_now[i][cabinet_num]))
    choice_room = []
    for el in range(len(head)):
        choice_room.append((str(transliterate_def(head[el])), str(head[el])))
    return(cosmetic_in_cab, choice_room)


def start_otchet(name, master, my_data):
    workbook_temp = openpyxl.load_workbook('shablon.xlsx')
    data = workbook_temp["Шаблон"]
    row = 7
    itog = 0
    otchet = []
    itog_brand = 0
    m = read_file('masters')
    n = read_file('head')
    for i in range(0, len(m)-1):
        if transliterate_def(m[i]) == master:
            master = m[i]
    for i in range(0, len(n)-1):
        if transliterate_def(n[i]) == name:
            name = n[i]
    data['B4'] = master
    data['B2'] = name
    try:
        name = transliterate_def(name)
    except:
        pass
    old_info, non = build_vir_cab(name, 1)
    temp = []
    for i in range(0, len(my_data)-1):
        temp.append(str(my_data[i][1]))
        if str(my_data[i][1]) != '':
            if float(old_info[i][3])-float(my_data[i][1])>=0:
                otchet.append([old_info[i][0], old_info[i][1], float(old_info[i][3])-float(my_data[i][1]), float(old_info[i][4])*(float(old_info[i][3])-float(my_data[i][1]))])
                print(old_info)
            else:
                error_weidth()

    x = len(otchet)
    for i in range(0, x):
        if i != 0 and otchet[i][0] != otchet[i - 1][0]:
            data["E" + str(row)] = itog_brand
            data["E" + str(row)].font = Font(bold=True, size=35)
            data["D" + str(row)] = "Итого"
            data["D" + str(row)].font = Font(bold=True, size=35)
            itog_brand = 0
            row = row + 2
        data["B" + str(row)] = otchet[i][0]
        data["C" + str(row)] = otchet[i][1]
        data["D" + str(row)] = otchet[i][2]
        data["E" + str(row)] = otchet[i][3]
        itog = float(data["E" + str(row)].value + itog)
        itog_brand = itog_brand + otchet[i][3]
        row = row + 1

    data["E" + str(row)] = itog_brand
    data["E" + str(row)].font = Font(bold=True, size=35)
    data["D" + str(row)] = "Итого"
    data["D" + str(row)].font = Font(bold=True, size=35)
    data["E" + str(row + 1)] = itog
    data["E" + str(row + 1)].font = Font(bold=True, size=35)
    data["D" + str(row + 1)] = "Итого расход"
    data["D" + str(row + 1)].font = Font(bold=True, size=35)
    dir_path = pathlib.Path.cwd()
    day = str(datetime.date.today() - datetime.timedelta(days=1))
    day = datetime.date.today()-datetime.timedelta(days=1)
    day = str(day.strftime("%d/%m/%Y"))
    data["B3"] = day
    try:
        os.makedirs('history' + "\\" + str(datetime.date.today()) + '')
    except:
        pass
    workbook_temp.save(
        pathlib.Path(dir_path, 'history' + "\\" + str(datetime.date.today()) + '', '' + name + '.xlsx'))


def write_in_excel(name, master, my_data):
    my_data = list(my_data.items())
    otchet_choise = False
    if my_data[1][1] == "on":
        otchet_choise = True
    del my_data[1]
    del my_data[0]
    for i in range(len(my_data)):
        my_data[i]=(my_data[0], my_data[i][1].replace(",", "."))
    if otchet_choise != False:
        start_otchet(name,master,my_data)
    cosmetic_in_cab, choice_room = build_vir_cab(name)
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp["Производство"]
    simvols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    coord = None
    for i in range(4, 1000):
        if transliterate_def(data[str(simvols[i]) + '1'].value) == name:
            coord = simvols[i]
            break
    for j in range(0, len(cosmetic_in_cab)):
        for i in range(1, 1000):
            if cosmetic_in_cab[j][1] == data['D'+str(i)].value:
                if str(my_data[j][1]) != '':
                    temp = data[coord + str(i)].value
                    data[coord + str(i)] = (my_data[j][1])
                    history_save(str(" в кабинете "+str(name)+" у мастера "+ master+" изменен вес " +str(data['D' + str(i)].value)+" c "+str(temp)+" на "+str(my_data[j][1])))
    workbook_temp.save('etual.xlsx')


def plus_in_exel(cabinet, product):
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp["Производство"]
    simvols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    coord = None
    for i in range(4, 1000):
        if transliterate_def(data[str(simvols[i]) + '1'].value) == cabinet:
            coord = simvols[i]
            break
    for j in range(0, len(product)):
        for i in range(1, 1000):
            if product[j][0] == data['D' + str(i)].value:
                if str(product[j][1]) != '':
                    try:
                        data[coord + str(i)] = float((product[j][2]).replace(",", ".")) + float(data[coord + str(i)].value)
                    except:
                        data[coord + str(i)] = float((product[j][2]).replace(",", "."))
                    history_save(" в кабинет "+str(cabinet)+" выдан продукт "+product[j][0])
    workbook_temp.save('etual.xlsx')


def cab_plus(request, p_choice_room, brands_views):
    class Plus_cab(forms.Form):
        brands = read_file("brands")
        brands_ch = []
        x = str(request.build_absolute_uri())
        r = x.replace('/', ' ')
        r = r.split()
        x = r[-1]
        c = r[-2]
        save = None
        cosm = []
        cosmetic_in_cab, choice_room = build_vir_cab("Наименование")
        for i in range(len(cosmetic_in_cab)):
            if transliterate_def(cosmetic_in_cab[i][0]) == x:
                cosm.append((i, cosmetic_in_cab[i][1], cosmetic_in_cab[i][2]))
        name_brands = None
        if len(x) == 3:
            name_brands = brands[0][0]

        for el in range(len(brands)):
            brands_ch.append((str(transliterate_def(brands[el][0])), str(brands[el][0])))
            if x==str(transliterate_def(brands[el][0])):
                name_brands = str(transliterate_def(brands[el][0]))
        brands_ch = list(set(brands_ch))
        brands_ch.sort()
        brands_ch_f = forms.ChoiceField(choices=brands_ch, label='', initial = x)

    if request.method == "POST":
        form = Plus_cab(request.POST, request.FILES)

        if form.is_valid():
            data = request.POST
            x = str(request.build_absolute_uri())
            r = x.replace('/', ' ')
            r = r.split()
            x = r[-1]
            c = r[-2]
            #запись добавления в ексель
            product = []
            for key in data:
                for num in Plus_cab.cosm:
                    if str(num[0]) == str(key) and str(data[key]) != '':
                        product.append((str(num[1]), str(num[2]), str(data[key])))
            plus_in_exel(c, product)
            info_base = form.cleaned_data
            return redirect(
                f"/plus/{c}/{info_base.get('brands_ch_f').lower()}")
    else:
        form = Plus_cab()
        return render(request, "control/cab_plus.html", {"form": form})


def cabinet_weight(request, cabinet, master):
    class Cabinet_weight(forms.Form):
        cosmetic_in_cab, choice_room = build_vir_cab(cabinet)
        for i in range(len(cosmetic_in_cab)):
            cosmetic_in_cab[i]=list(cosmetic_in_cab[i])
            cosmetic_in_cab[i].append(str(i))
        x = str(request.build_absolute_uri())
        x = x[x.rindex("/") + 1:]

    name = str(request.build_absolute_uri())
    name = name.replace("/", " ")
    name = name.split()
    master = str(name[-1])
    name = str(name[-2])


    if request.method == "POST":
        form = Cabinet_weight(request.POST, request.FILES)
        if form.is_valid():
            data = request.POST
            write_in_excel(name, master, data)
            return redirect(f"/", {"form": form})
    else:
        form = Cabinet_weight()
        return render(request, "control/cabinet_weight.html", {"form": form})


def weighing(request):
    if request.method == "POST":
        form = Weighing(request.POST, request.FILES)
        if form.is_valid():
            info_base = form.cleaned_data
            return redirect(f"/weighing/{info_base.get('p_choice_room').lower()}/{info_base.get('choice_master').lower()}")
    else:
        form = Weighing()
        return render(request, "control/weighing.html", {"form": form})


def home(request):
    data = read_file('all')
    return render(request, "control/home.html", {"data": data})
